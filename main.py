import pandas as pd
import numpy as np
import os
from openpyxl import load_workbook
from sklearn.experimental import enable_iterative_imputer
from sklearn.impute import IterativeImputer
from scipy.stats import zscore
from scipy.stats.mstats import winsorize

def cap_outliers_zscore(df, threshold=3):
    z_scores = np.abs(zscore(df, nan_policy='omit'))
    df[(z_scores > threshold)] = np.nan
    return df

def winsorize_data(df, limits=[0.05, 0.05]):
    return df.apply(lambda x: winsorize(x, limits=limits) if x.dtype.kind in 'biufc' else x, axis=0)

def process_excel_file(file_path, output_folder, stddev_folder, heatmap_folder):
    # Extract base name and extension for file naming
    base_name = os.path.basename(file_path)
    name, ext = os.path.splitext(base_name)

    # Read all sheets into a dictionary of DataFrames
    sheets_dict = pd.read_excel(file_path, sheet_name=None)
    std_devs = {}

    # Initialize the IterativeImputer with increased max_iter
    imputer = IterativeImputer(max_iter=100, random_state=0)

    # Iterate through each sheet and process the data
    for sheet_name, df in sheets_dict.items():
        # Replace 0 with NaN, except for the B2 cell
        df.replace(0, np.nan, inplace=True)
        df.iat[0, 1] = 0  # Restore B2 value

        # Remove columns that are all zeros except for the value in the first row
        cols_to_drop = [col for col in df.columns if df[col].iloc[0] != 0 and df[col].iloc[1:].isna().all()]
        df.drop(columns=cols_to_drop, inplace=True)

        # New logic to remove columns with 30% or more zeroes
        threshold = 0.3  # 30%
        for col in df.columns:
            zero_count = df[col].isna().sum()
            total_count = len(df[col])
            zero_percentage = zero_count / total_count
            if zero_percentage >= threshold:
                df.drop(columns=[col], inplace=True)

        # Convert column names to string
        df.columns = df.columns.astype(str)

        # Use IterativeImputer to fill missing values
        numeric_df = df.select_dtypes(include=[np.number])
        if not numeric_df.empty:
            imputed_data = imputer.fit_transform(numeric_df)
            df.loc[:, numeric_df.columns] = imputed_data

            # Handle outliers using Z-Score method
            capped_df = cap_outliers_zscore(pd.DataFrame(imputed_data, columns=numeric_df.columns))
            # Apply Winsorization
            winsorized_df = winsorize_data(capped_df)
            std_devs[sheet_name] = winsorized_df.std()

        # Update the dictionary with the imputed and winsorized DataFrame
        sheets_dict[sheet_name] = df

    # Generate the output file path with "CLEANED(EM)" suffix
    output_file_name = f"{name}_CLEANED(EM){ext}"
    output_file_path = os.path.join(output_folder, output_file_name)

    # Save the cleaned data back to an Excel file
    with pd.ExcelWriter(output_file_path, engine='openpyxl') as writer:
        for sheet_name, df in sheets_dict.items():
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    # Adjust column widths to fit the content
    wb = load_workbook(output_file_path)
    for sheet_name in wb.sheetnames:
        sheet = wb[sheet_name]
        for col in sheet.columns:
            max_length = 0
            column = col[0].column_letter  # Get the column name
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            sheet.column_dimensions[column].width = adjusted_width
    wb.save(output_file_path)

    # Ensure the stddev folder exists
    if not os.path.exists(stddev_folder):
        os.makedirs(stddev_folder)

def process_all_excel_files(input_folder, output_folder, stddev_folder, heatmap_folder):
    # Ensure the output folder exists
    if not os.path.exists(output_folder):
        os.makedirs(output_folder)

    # Ensure the standard deviations folder exists
    if not os.path.exists(stddev_folder):
        os.makedirs(stddev_folder)

    # Ensure the heatmap folder exists
    if not os.path.exists(heatmap_folder):
        os.makedirs(heatmap_folder)

    # Iterate over all .xlsx files in the input folder
    for file_name in os.listdir(input_folder):
        if file_name.endswith('.xlsx'):
            file_path = os.path.join(input_folder, file_name)
            process_excel_file(file_path, output_folder, stddev_folder, heatmap_folder)

def main():
    input_folder = 'data'
    output_folder = 'cleaned_data'
    stddev_folder = 'standarddeviations'
    heatmap_folder = 'heatmaps'

    # Process all files in the input folder
    process_all_excel_files(input_folder, output_folder, stddev_folder, heatmap_folder)

if __name__ == '__main__':
    main()
